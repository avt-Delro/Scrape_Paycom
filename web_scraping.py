
from playwright.sync_api import sync_playwright
import environment as env
from dotenv import load_dotenv
import win32com.client as win32
import os
from datetime import datetime
import calendar
import pandas as pd
from openpyxl import load_workbook




load_dotenv()
paycom_user = env.paycom_username
paycom_pass = env.paycom_password  
client_code = env.paycom_clientcode 

local_path = env.paycom_local
email = env.sendemail

datetoday = datetime.today()



outlook = win32.Dispatch("Outlook.Application")
outlook_ap = outlook.GetNamespace("MAPI")

def paycom_scraping(weblink, username, password, client_code):
    try:
        with sync_playwright() as p:
            date = datetoday.strftime("%m/%d/%Y")
            
            
            context = p.chromium.launch_persistent_context(
                    user_data_dir="edge_automation_profile",
                    channel="msedge",
                    headless=True
                )


            page = context.new_page()
            
            page.goto(weblink, wait_until='load') 
            page.locator("#clientcode").fill(client_code)
            page.locator("#username").fill(username)
            page.locator("#password").fill(password)
            page.get_by_role("button", name="Log In").click()
            page.wait_for_load_state("networkidle")
            page.goto("https://www.paycomonline.net/v4/cl/rpt-center.php",wait_until='networkidle')
            page.get_by_role("tab", name="Push Reporting™").click()
            page.get_by_role("row", name=f"{date} Actual v Scheduled").get_by_label("More options dropdown").click()
            page.get_by_role("link", name="View Files").click()
            with page.expect_download() as download_info:
                page.get_by_role("link", name="Download").click()
                download = download_info.value

            file_folder = os.path.join(local_path, 'HR Files')
            os.makedirs(file_folder, exist_ok=True)

            file_name =  f'OT_report_{datetoday.strftime("%Y%m%d")}.xlsx'

            download.save_as(os.path.join(file_folder,file_name))

            return f"{str(file_folder)}\{file_name}"

    except Exception as e:
        print(e)

def create_sheet(filepath, data_row, sheetname):
    if isinstance(data_row, list):
        df = pd.DataFrame(data_row)
    elif isinstance(data_row, dict):
        df = pd.DataFrame([data_row])
    
    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer,  sheet_name=sheetname, index=False)


def create_report(file):
    df = pd.read_excel(file)

    cols = ["Scheduled Hours", "Actual Hours", "Variance"]
    df[cols] = df[cols].apply(pd.to_numeric, errors="coerce").fillna(0)

    df['Punch Date'] = pd.to_datetime(df['Punch Date'], errors='coerce')

    df_filtered = df[df['Punch Date']<= datetoday.strftime('%m/%d/%Y')]


    summary = (
        df_filtered.groupby("Employee", as_index=False)
        .agg({
            "Scheduled Hours": "sum",
            "Actual Hours": "sum",
            "Variance": "sum",
        })
    )
    create_sheet(file, summary.to_dict(orient="records"), "Summary")

def send_email (email, filepath):
    outlook = win32.Dispatch("Outlook.Application")
    outlook_ap = outlook.GetNamespace("MAPI")
    mail = outlook.CreateItem(0)

    wb = load_workbook(filepath)
    wb._sheets = [wb[s] for s in ["Summary","Actual vs. Schedu"]]
    wb.save(filepath)

    mail.Attachments.Add(filepath)

    mail.To = email
    mail.Subject = f'See attached file, for Paycom Data for the date: {datetoday.strftime("%m/%d/%Y")}'
    mail.Send()
    print('Email Sent')
    os.remove(filepath)



paycom_filepath = paycom_scraping('https://www.paycomonline.net/v4/cl/cl-login.php', paycom_user, paycom_pass, client_code)
create_report(paycom_filepath)
send_email(email, paycom_filepath)